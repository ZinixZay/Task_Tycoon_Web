from random import randint
import json
from transliterate import translit

from django.contrib.auth.mixins import AccessMixin
from django.utils.text import slugify

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.cell import get_column_letter

# Variables

menu = [
    {'title': 'Главная', 'url_name': 'home'},
    {'title': 'Мои задания', 'url_name': 'tasks'},
    {'title': 'Создать задание', 'url_name': 'createtask'},
    {'title': 'Решить задание', 'url_name': 'search'}
]


# Mixins


class DataMixin:
    """
    Collecting args and adding menu to a page
    """
    def set_context(self, **kwargs):
        context = kwargs
        context['menu'] = menu
        return context


class AuthorRequiredMixin(AccessMixin):
    """
    Permission mixin only for object author
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        if request.user != self.get_object().creator:
            return self.handle_no_permission()
        return super().dispatch(request, *args, **kwargs)


# Functions


def generate_slug(title: str, task_model) -> str:
    """
    Generating slug for new task
    :param title: task's title
    :param task_model: Task model
    :return: generated slug
    """
    slug = slugify(title, allow_unicode=True)
    slug = translit(slug, 'ru', reversed=True)
    slug = slug.replace("'", "")
    similar_slugs = sorted([i.slug for i in task_model.objects.all() if slug in i.slug])
    match len(similar_slugs):
        case 0:
            return slug
        case 1:
            return slug + '1'
    slug = similar_slugs[-1]
    return slug[:len(slug) - 1] + (str(int(slug[-1]) + 1))


def generate_identifier() -> int:
    """
    Generating random identifier for task. Checks if it is unique
    :return: identifier
    """
    from .models import Task
    while True:
        identifier = randint(100000, 999999)
        if identifier not in [i.identifier for i in Task.objects.all()]:
            break
    return identifier


def parse_answer_to_dict(response) -> dict:
    """
    Parsing user's answer to a python dict
    :param response: user's task answer
    :return: parsed user's answer
    """
    str_response = str(response)
    str_response = str_response.replace("'", '"')
    dict_response = json.loads(str_response[12:len(str_response) - 1])
    content = {}
    task_title = None
    for header in dict_response:
        if header == 'csrfmiddlewaretoken':
            continue
        if header == 'task_title':
            task_title = dict_response[header]
            continue
        content[header] = dict_response[header]
    return {'task_title': task_title[0], **content}


def check_solution_allowed(Answer, task, user) -> bool:
    """
    Check if user already answered this task. if yes -> does not accept this solution
    :param Answer: Answer model
    :param task: solution task
    :param user: solution user
    :return: True, if solution exists
    """
    if len(Answer.objects.filter(task=task, user=user)) == task.attempts:
        return False
    return True


def analyse_answer(question_query, user_answers) -> dict:
    """
    Gets user's solution, right solution from data and compare them. Forming a dict with right and false answers
    :param question_query: All in-task questions
    :param user_answers: User's solution
    :return: python dict with rightness of answers
    """
    right_answers = dict()
    for question in question_query:
        if question.test_type:
            right_answers[question.title] = [i['response_name'] for i in list(
                filter(lambda x: x['response_right'], question.variants))]
        else:
            right_answers[question.title] = question.variants.lower()

    result = dict()
    for question, answer in user_answers.content.items():
        if question in right_answers.keys():
            if type(right_answers[question]) is str:
                if answer[0].lower() == right_answers[question]:
                    result[question] = True
                else:
                    result[question] = False
            else:
                if answer != right_answers[question]:
                    result[question] = False
                else:
                    result[question] = True
    return result


def generate_excel(task, answers, questions):
    wb = Workbook()
    ws = wb.active

    ws.title = task.title
    ws["A1"] = "№"
    ws["B1"] = "Ученик"
    ws.column_dimensions["B"].width = 25

    curr_column = 3
    curr_row = 1

    for question in questions:
        ws.column_dimensions[get_column_letter(curr_column)].width = len(question.title) + 2
        ws.cell(row=curr_row, column=curr_column, value=question.title)
        curr_column += 1
    for col in ws.iter_cols(min_row=1, max_row=len(set([i.user_id for i in answers])) + 1,
                            max_col=len(questions) + 4, min_col=1):
        for cell in col:
            cell.alignment = Alignment(horizontal='center')

    ws.cell(row=curr_row, column=curr_column, value='Кол-во верных ответов')
    ws.column_dimensions[get_column_letter(curr_column)].width = 25

    ws.cell(row=curr_row, column=curr_column + 1, value='Кол-во попыток')
    ws.column_dimensions[get_column_letter(curr_column + 1)].width = 17

    curr_row += 1
    curr_column = 1

    answered_persons = sorted(list(set(i.user.username for i in answers)))

    for user_id in range(len(answered_persons)):
        ws.cell(row=curr_row, column=curr_column, value=user_id + 1)
        ws.cell(row=curr_row, column=curr_column + 1, value=answered_persons[user_id])

        curr_column = 3
        for i in range(len(questions)):
            for answer in list(filter(lambda ans: ans.user.username == answered_persons[user_id], answers)):
                if type(questions[i].variants) is str:  # Если не тестовый тип
                    if answer.content[questions[i].title] != ['']:
                        if answer.content[questions[i].title][0].lower() == questions[i].variants.lower():
                            ws.cell(row=curr_row, column=curr_column, value='+++')
                        else:
                            if next(ws.iter_cols(min_row=curr_row, max_row=curr_row, min_col=curr_column,
                                                 max_col=curr_column))[0].value != '+++':
                                ws.cell(row=curr_row, column=curr_column, value='---')
                else:
                    if str(questions[i]) in answer.content.keys():
                        true_answers = [j['response_name'] for j in questions[i].variants if j['response_right']]
                        if answer.content[str(questions[i])] == true_answers:
                            ws.cell(row=curr_row, column=curr_column, value='+++')
                        else:
                            if next(ws.iter_cols(min_row=curr_row, max_row=curr_row, min_col=curr_column,
                                                 max_col=curr_column))[0].value != '+++':
                                ws.cell(row=curr_row, column=curr_column, value='---')
            curr_column += 1

        t_answers = 0
        for col in ws.iter_cols(min_col=3, max_col=curr_column - 1, min_row=curr_row, max_row=curr_row):
            for cell in col:
                if cell.value == '+++':
                    cell.fill = PatternFill('solid', fgColor="ACB78E")
                    t_answers += 1
                elif cell.value == '---':
                    cell.fill = PatternFill('solid', fgColor="B17267")
                else:
                    cell.fill = PatternFill('solid', fgColor="DCDCDC")
        ws.cell(row=curr_row, column=curr_column, value=t_answers)
        ws.cell(row=curr_row, column=curr_column + 1, value=len(list(filter(
            lambda ans: ans.user.username == answered_persons[user_id], answers))))

        curr_column = 1
        curr_row += 1

    path_to_save = f"media/uploads/{task.title}.xlsx"
    wb.save(path_to_save)
    return path_to_save
